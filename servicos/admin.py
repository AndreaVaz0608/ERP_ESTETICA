from django.contrib import admin
from django.http import HttpResponse
from .models import Servico
import openpyxl

@admin.register(Servico)
class ServicoAdmin(admin.ModelAdmin):
    list_display = ('nome', 'preco', 'duracao_minutos', 'ativo')
    search_fields = ('nome',)
    list_filter = ('ativo',)
    actions = ['exportar_para_excel']

    @admin.action(description="Exportar serviços selecionados para Excel")
    def exportar_para_excel(self, request, queryset):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Serviços"

        # Cabeçalho
        ws.append(['Nome', 'Preço (R$)', 'Duração (min)', 'Ativo'])

        # Dados
        for servico in queryset:
            row = [
                servico.nome,
                float(servico.preco),
                servico.duracao_minutos,
                'Sim' if servico.ativo else 'Não'
            ]
            ws.append(row)

            # Formatar o valor como moeda R$ na célula de preço (coluna 2)
            preco_cell = ws.cell(row=ws.max_row, column=2)
            preco_cell.number_format = 'R$ #,##0.00'  # Formato manual para reais

        # Geração do arquivo de resposta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=servicos_silvia_vaz.xlsx'
        wb.save(response)
        return response
